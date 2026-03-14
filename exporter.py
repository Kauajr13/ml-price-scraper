import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from scraper import Product
from analyzer import PriceAnalyzer


DARK = "1A1A2E"
LIGHT_TEXT = "E0E0E0"
YELLOW = "FFE600"
ROW_ALT = "F5F8FF"
SECTION_BG = "EEF2FF"
BORDER_COLOR = "CCCCCC"


def thin_border():
    s = Side(style="thin", color=BORDER_COLOR)
    return Border(left=s, right=s, top=s, bottom=s)


def dark_fill():
    return PatternFill("solid", fgColor=DARK)


def alt_fill():
    return PatternFill("solid", fgColor=ROW_ALT)


def section_fill():
    return PatternFill("solid", fgColor=SECTION_BG)


def write_products_sheet(wb: openpyxl.Workbook, products: list[Product]):
    ws = wb.active
    ws.title = "Produtos"
    ws.sheet_view.showGridLines = False

    columns = [
        ("Titulo", 55),
        ("Preco (R$)", 14),
        ("Preco Original (R$)", 20),
        ("Desconto (%)", 14),
        ("Rating", 10),
        ("Reviews", 12),
        ("Vendidos", 18),
        ("Frete Gratis", 14),
        ("Condicao", 12),
        ("Vendedor", 25),
        ("URL", 60),
        ("Coletado em", 20),
    ]

    ws.merge_cells("A1:L1")
    title = ws["A1"]
    title.value = "ML Price Scraper"
    title.font = Font(name="Calibri", bold=True, size=15, color=YELLOW)
    title.fill = dark_fill()
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    for col_idx, (label, width) in enumerate(columns, start=1):
        cell = ws.cell(row=2, column=col_idx, value=label)
        cell.font = Font(name="Calibri", bold=True, size=10, color=LIGHT_TEXT)
        cell.fill = dark_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[2].height = 22

    for row_idx, p in enumerate(products, start=3):
        row_data = [
            p.title,
            p.price,
            p.original_price or "",
            p.discount_pct or "",
            p.rating or "",
            p.reviews_count or "",
            p.sold_count or "",
            "Sim" if p.shipping_free else "Nao",
            p.condition,
            p.seller or "",
            p.url,
            p.scraped_at,
        ]
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name="Calibri", size=9)
            cell.border = thin_border()
            cell.alignment = Alignment(vertical="center", wrap_text=(col_idx == 1))

            if col_idx == 2 and isinstance(value, float):
                cell.number_format = 'R$ #,##0.00'
            if col_idx == 3 and isinstance(value, float):
                cell.number_format = 'R$ #,##0.00'
            if col_idx == 4 and isinstance(value, float):
                cell.number_format = '0.0"%"'

            if col_idx == 8:
                bg = "D4EDDA" if p.shipping_free else "F8D7DA"
                cell.fill = PatternFill("solid", fgColor=bg)
            elif col_idx == 4 and isinstance(value, float) and value >= 20:
                cell.fill = PatternFill("solid", fgColor="FFF3CD")
            elif row_idx % 2 == 0:
                cell.fill = alt_fill()

        ws.row_dimensions[row_idx].height = 18

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:L{len(products) + 2}"


def write_summary_sheet(wb: openpyxl.Workbook, analyzer: PriceAnalyzer, keyword: str):
    ws = wb.create_sheet("Resumo")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 22

    def section_header(row, title):
        ws.merge_cells(f"A{row}:B{row}")
        c = ws.cell(row=row, column=1, value=f"  {title}")
        c.font = Font(name="Calibri", bold=True, size=11, color=LIGHT_TEXT)
        c.fill = dark_fill()
        c.alignment = Alignment(vertical="center")
        ws.row_dimensions[row].height = 22

    def row_kv(row, key, value, fmt=None):
        kc = ws.cell(row=row, column=1, value=key)
        vc = ws.cell(row=row, column=2, value=value)
        for c in (kc, vc):
            c.font = Font(name="Calibri", size=10)
            c.fill = section_fill()
            c.border = thin_border()
        vc.font = Font(name="Calibri", bold=True, size=10)
        vc.alignment = Alignment(horizontal="center")
        if fmt:
            vc.number_format = fmt
        ws.row_dimensions[row].height = 18

    s = analyzer.summary()

    ws.merge_cells("A1:B1")
    t = ws["A1"]
    t.value = f"Analise: {keyword.title()}"
    t.font = Font(name="Calibri", bold=True, size=14, color=YELLOW)
    t.fill = dark_fill()
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    section_header(2, "Precos")
    row_kv(3,  "Total de produtos",   s.get("total", 0))
    row_kv(4,  "Preco minimo",        s.get("price_min", 0),    "R$ #,##0.00")
    row_kv(5,  "Preco maximo",        s.get("price_max", 0),    "R$ #,##0.00")
    row_kv(6,  "Preco medio",         s.get("price_avg", 0),    "R$ #,##0.00")
    row_kv(7,  "Mediana",             s.get("price_median", 0), "R$ #,##0.00")
    row_kv(8,  "Desvio padrao",       s.get("price_stdev", 0),  "R$ #,##0.00")

    section_header(10, "Frete e Descontos")
    row_kv(11, "Com frete gratis",    s.get("free_shipping_count", 0))
    row_kv(12, "% frete gratis",      s.get("free_shipping_pct", 0), '0.0"%"')
    row_kv(13, "Com desconto",        s.get("discounted_count", 0))
    row_kv(14, "Desconto medio",      s.get("avg_discount_pct", 0), '0.0"%"')

    section_header(16, "Avaliacoes")
    row_kv(17, "Com avaliacao",       s.get("rated_count", 0))
    row_kv(18, "Rating medio",        s.get("avg_rating") or "N/A")

    section_header(20, "Distribuicao de Precos")
    ws.cell(row=21, column=1, value="Faixa").font = Font(bold=True)
    ws.cell(row=21, column=2, value="Quantidade").font = Font(bold=True)

    ranges = analyzer.price_ranges()
    for i, (label, count) in enumerate(ranges.items(), start=22):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=count)

    if len(ranges) >= 2:
        chart = BarChart()
        chart.type = "col"
        chart.title = "Distribuicao de Precos"
        chart.y_axis.title = "Produtos"
        chart.x_axis.title = "Faixa de Preco"
        chart.style = 10
        chart.width = 18
        chart.height = 10

        data_ref = Reference(ws, min_col=2, min_row=21, max_row=21 + len(ranges))
        cats_ref = Reference(ws, min_col=1, min_row=22, max_row=21 + len(ranges))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        ws.add_chart(chart, "D3")


def write_deals_sheet(wb: openpyxl.Workbook, analyzer: PriceAnalyzer):
    ws = wb.create_sheet("Top Deals")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value = "Melhores Oportunidades"
    t.font = Font(name="Calibri", bold=True, size=13, color=YELLOW)
    t.fill = dark_fill()
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = ["#", "Titulo", "Preco (R$)", "Desconto (%)", "Rating", "Frete Gratis"]
    widths =  [5,   55,        14,            14,             10,        14]

    for col, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = Font(name="Calibri", bold=True, size=10, color=LIGHT_TEXT)
        c.fill = dark_fill()
        c.alignment = Alignment(horizontal="center")
        c.border = thin_border()
        ws.column_dimensions[get_column_letter(col)].width = w

    medal_colors = {1: "FFF3CD", 2: "E8E8E8", 3: "FDEBD0"}

    for i, p in enumerate(analyzer.top_deals(10), start=1):
        row = i + 2
        row_data = [
            i, p.title, p.price,
            p.discount_pct or "",
            p.rating or "",
            "Sim" if p.shipping_free else "Nao",
        ]
        for col, v in enumerate(row_data, start=1):
            c = ws.cell(row=row, column=col, value=v)
            c.font = Font(name="Calibri", size=9)
            c.border = thin_border()
            c.alignment = Alignment(
                horizontal="left" if col == 2 else "center",
                vertical="center"
            )
            if col == 3:
                c.number_format = "R$ #,##0.00"
            if i in medal_colors:
                c.fill = PatternFill("solid", fgColor=medal_colors[i])
            elif row % 2 == 0:
                c.fill = alt_fill()
        ws.row_dimensions[row].height = 18


def export(
    products: list[Product],
    analyzer: PriceAnalyzer,
    keyword: str,
    output_dir: str = "output"
) -> str:
    os.makedirs(output_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    slug = keyword.replace(" ", "_").lower()
    filename = f"{output_dir}/ml_{slug}_{timestamp}.xlsx"

    wb = openpyxl.Workbook()
    write_products_sheet(wb, products)
    write_summary_sheet(wb, analyzer, keyword)
    write_deals_sheet(wb, analyzer)
    wb.save(filename)

    return filename
